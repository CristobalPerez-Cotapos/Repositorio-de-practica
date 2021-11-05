from gurobipy import GRB, Model, quicksum
from openpyxl import load_workbook
import parametros as p

model = Model("Distribución de tercera dosis")

ruta_datos = "Datos este si que si.xlsx"
libro_datos = load_workbook(ruta_datos)

hojas = libro_datos.get_sheet_names()

letras = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

## Primero cargamos los parametros

diccionario_parametros = {}

for hoja in hojas:
    sheet = libro_datos[hoja]
    filas = sheet.max_row
    columnas = sheet.max_column
    diccionario_parametros[hoja] = {}
    for i in range(columnas):
        diccionario_parametros[hoja][i+1] = {}
        for j in range(filas):
            diccionario_parametros[hoja][i+1][j+1] = sheet[letras[i] + str(j+1)].value

## Luego definimos las variables


R_ijtm = {}
for i in range(p.I_CENTROS_VAC):
    for j in range(p.J_TIPOS_VAC):
        for t in range(p.T_DIAS):
            for m in range(p.M_METODOS_TRANSPORTE):
                R_ijtm[(i+1, j+1, t+1, m+1)] = model.addVar(
                    vtype=GRB.INTEGER, name=f"R_{str(i+1)}{str(j+1)}{str(t+1)}{str(m+1)}")
                model.addConstr(R_ijtm[(i+1, j+1, t+1, m+1)] >= 0, name="Naturaleza R")

CT_mt = {}
for m in range(p.M_METODOS_TRANSPORTE):
    for t in range(p.T_DIAS):
        CT_mt[(m+1, t+1)] = model.addVar(vtype=GRB.INTEGER, name=f"CT_{str(m+1)}{str(t+1)}")
        model.addConstr(CT_mt[(m + 1, t + 1)] >= 0, name="Naturaleza CT")

CC_imt = {}
for i in range(p.I_CENTROS_VAC):
    for m in range(p.M_METODOS_TRANSPORTE):
        for t in range(p.T_DIAS):
            CC_imt[(i+1, m+1, t+1)] = model.addVar(
                vtype=GRB.INTEGER, name=f"CC_{str(i+1)}{str(m+1)}{str(t+1)}")
            model.addConstr(CC_imt[(i+1, m+1, t+1)]  >= 0, name="Naturaleza CC")

CN_mt = {}
for m in range(p.M_METODOS_TRANSPORTE):
    for t in range(p.T_DIAS):
        CN_mt[(m+1, t+1)] = model.addVar(vtype=GRB.INTEGER, name=f"CN_{str(m+1)}{str(t+1)}")
        model.addConstr(CN_mt[(m+1, t+1)] >= 0, name="Naturaleza CN")

CA_met = {}
for m in range(p.M_METODOS_TRANSPORTE):
    for e in range(p.E_EMPRESAS_CAM):
        for t in range(p.T_DIAS):
            CA_met[(m+1, e+1, t+1)] = model.addVar(
                vtype=GRB.INTEGER, name=f"CA_{str(m+1)}{str(e+1)}{str(t+1)}")
            model.addConstr(CA_met[(m+1, e+1, t+1)] >= 0, name="Naturaleza CA")

I_tj = {}
for t in range(p.T_DIAS):
    for j in range(p.J_TIPOS_VAC):
        I_tj[(t+1, j+1)] = model.addVar(vtype=GRB.INTEGER, name=f"I_{str(t+1)}{str(j+1)}")
        model.addConstr(I_tj[(t+1, j+1)] >= 0, name="Naturaleza I")

E_et = {}
for e in range(p.E_EMPRESAS_CAM):
    for t in range(p.T_DIAS):
        E_et[(e+1, t+1)] = model.addVar(vtype=GRB.BINARY, name=f"E_{str(e+1)}{str(t+1)}")

DES_et = {}
for e in range(p.E_EMPRESAS_CAM):
    for t in range(p.T_DIAS):
        DES_et[(e+1, t+1)] = model.addVar(vtype=GRB.BINARY, name=f"DES_{str(e+1)}{str(t+1)}")

VPA_t = {}
for t in range(p.T_DIAS):
    VPA_t[(t+1)] = model.addVar(vtype=GRB.INTEGER, name=f"VPA_{str(t+1)}")
    model.addConstr(VPA_t[(t+1)] >= 0, name="Naturaleza VP")

VP_t = {}
for t in range(p.T_DIAS):
    VP_t[(t+1)] = model.addVar(vtype=GRB.CONTINUOUS, name=f"VP_{str(t+1)}")
    model.addConstr(VP_t[(t+1)] >= 0, name="Naturaleza VP")

model.update()

## Después declaramos las restricciones



for i in range(p.I_CENTROS_VAC):
    for m in range(p.M_METODOS_TRANSPORTE):
        for t in range(p.T_DIAS):
            model.addConstr(quicksum(R_ijtm[(i +1, j+1 , t +1 , m+1)] * diccionario_parametros["VOL_j"][j+1][1] for j in range(p.J_TIPOS_VAC))
                            <= diccionario_parametros["CAP_m"][m+1][1] * CC_imt[(i+1, m+1, t+1)],
                            name="R1")

for i in range(p.I_CENTROS_VAC):
    for j in range(p.J_TIPOS_VAC):
        for t in range(p.T_DIAS):
            model.addConstr(diccionario_parametros[f"DE_{str(i+1)}jt"][j+1][t+1] <=
                            quicksum(R_ijtm[(i+1, j+1, t+1, m+1)] * diccionario_parametros["CAC_j"][j+1][1]
                                     * diccionario_parametros["PPC_m"][m+1][1] for m in range(p.M_METODOS_TRANSPORTE)),
                            name= "R2")

for i in range(p.I_CENTROS_VAC):
    for t in range(p.T_DIAS):
        model.addConstr(quicksum(diccionario_parametros[f"DE_{str(i+1)}jt"][j+1][t+1] for j in range(p.J_TIPOS_VAC))
                        <= quicksum(R_ijtm[(i+1, j+1, t+1, m+1)] * diccionario_parametros["CAC_j"][j+1][1]
                                     * diccionario_parametros["PPC_m"][m+1][1] for j in range(p.J_TIPOS_VAC)
                                    for m in range(p.M_METODOS_TRANSPORTE)),
                        name="R3")

for j in range(p.J_TIPOS_VAC):
    for t in range(1, p.T_DIAS):
        model.addConstr(diccionario_parametros["V_jt"][j+1][t+1] + I_tj[(t, j+1)] ==
                        quicksum(R_ijtm[(i +1, j+1 , t +1, m+1)] for i in range(p.I_CENTROS_VAC) for m in range(p.M_METODOS_TRANSPORTE)) + I_tj[(t+1, j+1)]
                        , name="R4")

for j in range(p.J_TIPOS_VAC):
    for t in range(1):
        model.addConstr(diccionario_parametros["V_jt"][j+1][t+1] + diccionario_parametros["a_j"][j+1][1] ==
                        quicksum(R_ijtm[(i +1, j+1 , t +1, m+1)] for i in range(p.I_CENTROS_VAC) for m in range(p.M_METODOS_TRANSPORTE)) + I_tj[(t+1, j+1)]
                      , name="Caso limite de inventario")


for t in range(1, p.T_DIAS):
    model.addConstr(
        VP_t[t+1] == quicksum(quicksum(quicksum(R_ijtm[(i +1, j+1 , t +1, m+1)] * diccionario_parametros["CAC_j"][j+1][1] * diccionario_parametros["PPC_m"][m+1][1] for m in range(p.M_METODOS_TRANSPORTE))
                                       - diccionario_parametros[f"DE_{str(i+1)}jt"][j+1][t+1] for j in range(p.J_TIPOS_VAC)) - diccionario_parametros["DG_it"][i+1][t+1] for i in range(p.I_CENTROS_VAC))
                               + quicksum(R_ijtm[(i +1, j+1 , t +1, m+1)] * diccionario_parametros["CAC_j"][j+1][1] * (0) for m in range(p.M_METODOS_TRANSPORTE) for j in range(p.J_TIPOS_VAC) for i in range(p.I_CENTROS_VAC)),
    name = "R5")

for t in range(1):
    model.addConstr(
        VP_t[t+1] == quicksum(quicksum(quicksum(R_ijtm[(i +1, j+1 , t +1, m+1)] * diccionario_parametros["CAC_j"][j+1][1] for m in range(p.M_METODOS_TRANSPORTE))
                                    - diccionario_parametros[f"DE_{str(i+1)}jt"][j+1][t+1] for j in range(p.J_TIPOS_VAC)) - diccionario_parametros["DG_it"][i+1][t+1] for i in range(p.I_CENTROS_VAC))
                            + quicksum(R_ijtm[(i +1, j+1 , t +1, m+1)] * diccionario_parametros["CAC_j"][j+1][1] * (1 - diccionario_parametros["PPC_m"][m+1][1]) for m in range(p.M_METODOS_TRANSPORTE) for j in range(p.J_TIPOS_VAC) for i in range(p.I_CENTROS_VAC)), name = "Caso limite vacunas perdidas")

for t in range(p.T_DIAS):
    for m in range(p.M_METODOS_TRANSPORTE):
        model.addConstr(quicksum(CC_imt[(i+1, m+1, t+1)] for i in range(p.I_CENTROS_VAC)) <= CT_mt[(m+1, t+1)] + quicksum(CA_met[(m+1, e+1, t+1)] for e in range(p.E_EMPRESAS_CAM)),
       name= "R6")


for m in range(p.M_METODOS_TRANSPORTE):
    for t in range(1,p.T_DIAS):
        model.addConstr(CT_mt[(m+1, t+1)]==diccionario_parametros["CIN_m"][m+1][1] + quicksum(CN_mt[(m+1,k+1)] for k in range(t)), name = "R7")

for m in range(p.M_METODOS_TRANSPORTE):
    for t in range(1):
        model.addConstr(CT_mt[(m+1, t+1)] == diccionario_parametros["CIN_m"][m+1][1], name="Caso limite camiones")

for m in range(p.M_METODOS_TRANSPORTE):
    for t in range(p.T_DIAS):
        model.addConstr(CN_mt[(m+1, t+1)] <= diccionario_parametros["CD_mt"][m+1][t+1], name="R8")

for m in range(p.M_METODOS_TRANSPORTE):
    for e in range(p.E_EMPRESAS_CAM):
        for t in range(p.T_DIAS):
            model.addConstr(CA_met[(m+1, e+1, t+1)] <= diccionario_parametros[f"CAM_m{str(e+1)}t"][m+1][t+1]*E_et[(e+1, t+1)], name="R9")


for e in range(p.E_EMPRESAS_CAM):
    for t in range(p.T_DIAS):
        if t > 8:
            model.addConstr(quicksum(E_et[(e+1, k+1)] for k in range(t-7, t)) <=
                            6 + (t+1) * DES_et[(e+1, t+1)],
                            name="R10")

for t in range(p.T_DIAS):
    model.addConstr(quicksum(CT_mt[(m+1, t+1)] for m in range(p.M_METODOS_TRANSPORTE)) <=
                    diccionario_parametros["MES"][1][1],
                    name="R11")

for t in range(p.T_DIAS):
    model.addConstr(VP_t[(t+1)] <= VPA_t[(t+1)], name="Correción VP")
    model.addConstr(VP_t[(t + 1)] >= VPA_t[(t + 1)] - 1, name="Correción VP")




## Función objetivo

objetivo = quicksum(
    (quicksum(I_tj[(t+1,j+1)] * diccionario_parametros["CI_j"][j+1][1] for j in range(p.J_TIPOS_VAC)))
    +(quicksum(E_et[(e+1, t+1)] * diccionario_parametros["CE_e"][e+1][1] - DES_et[(e+1, t+1)] * diccionario_parametros["CED_e"][e+1][1] for e in range(p.E_EMPRESAS_CAM)))
    + quicksum(
        diccionario_parametros["CMP_m"][m+1][1] * CT_mt[(m+1,t+1)] + CN_mt[(m+1,t+1)] * diccionario_parametros["CCN_mt"][m+1][t+1] +
        quicksum(CA_met[(m+1, e+1, t+1)] * diccionario_parametros["CDA_me"][m+1][e+1] for e in range(p.E_EMPRESAS_CAM))
        + (quicksum(CC_imt[(i+1, m+1, t+1)] for i in range(p.I_CENTROS_VAC)) - quicksum(CA_met[(m+1,e+1, t+1)] for e in range(p.E_EMPRESAS_CAM))) * diccionario_parametros["CUP_m"][m+1][1]
        for m in range(p.M_METODOS_TRANSPORTE))
    + quicksum(R_ijtm[(i+1, j+1, t+1, m+1)] * diccionario_parametros[f"CTR_{str(i+1)}jm"][j+1][m+1] for i in range(p.I_CENTROS_VAC) for j in range(p.J_TIPOS_VAC) for m in range(p.M_METODOS_TRANSPORTE)) +
    VPA_t[(t+1)] * diccionario_parametros["CPV"][1][1] for t in range(p.T_DIAS))


model.update()
## Finalmente

model.setObjective(objetivo, GRB.MINIMIZE)

model.update()

model.optimize()



model.printAttr("x")

